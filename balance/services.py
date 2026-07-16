"""
services.py - لایه منطق کسب‌وکار برای سیستم موازنه متریال جهانپارس
=======================================================================
این فایل مسئول پردازش داده‌ها، اجرای فرمول‌های موازنه و تولید خروجی اکسل است.
ساختار خروجی:
  - یک شیت به ازای هر پیمانکار
  - بالای هر شیت: اطلاعات پیمانکار، شماره قرارداد، موضوع قرارداد، بازه زمانی
  - جدول داده‌ها با قابلیت فیلتر (AutoFilter)
"""

import io
from decimal import Decimal, ROUND_HALF_UP
import datetime
import jdatetime

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from django.db.models import Sum, Min, Max


# ─────────────────────────────────────────────────────────────────────────────
# ثابت‌های رنگ برای قالب‌بندی اکسل
# ─────────────────────────────────────────────────────────────────────────────
COLOR_HEADER_BG      = "1F3864"   # آبی تیره - هدر اصلی
COLOR_HEADER_FONT    = "FFFFFF"   # سفید
COLOR_SUB_HEADER_BG  = "2E75B6"   # آبی متوسط - ردیف عناوین ستون
COLOR_INFO_BG        = "EBF3FB"   # آبی خیلی روشن - بلوک اطلاعات پیمانکار
COLOR_INFO_LABEL_BG  = "D6E4F0"   # کمی تیره‌تر برای برچسب‌ها
COLOR_ROW_ODD        = "DCE6F1"   # آبی خیلی روشن - ردیف‌های فرد
COLOR_ROW_EVEN       = "FFFFFF"   # سفید - ردیف‌های زوج
COLOR_POSITIVE       = "C6EFCE"   # سبز روشن - مازاد پرداخت
COLOR_POSITIVE_FONT  = "276221"   # سبز تیره
COLOR_NEGATIVE       = "FFC7CE"   # قرمز روشن - کسری متریال
COLOR_NEGATIVE_FONT  = "9C0006"   # قرمز تیره
COLOR_ZERO           = "FFEB9C"   # زرد روشن - ایده‌آل
COLOR_ZERO_FONT      = "9C6500"   # نارنجی تیره
COLOR_BORDER         = "BDD7EE"   # آبی کم‌رنگ برای خطوط جدول
COLOR_INFO_BORDER    = "AEC6CF"   # رنگ حاشیه بلوک اطلاعات


# ─────────────────────────────────────────────────────────────────────────────
# کش استایل‌های OpenPyXL (افزایش بسیار زیاد سرعت تولید فایل)
# ─────────────────────────────────────────────────────────────────────────────
GLOBAL_BORDER = Border(
    left=Side(style='thin', color=COLOR_BORDER),
    right=Side(style='thin', color=COLOR_BORDER),
    top=Side(style='thin', color=COLOR_BORDER),
    bottom=Side(style='thin', color=COLOR_BORDER)
)

GLOBAL_ALIGNMENT_RIGHT = Alignment(horizontal='right', vertical='center', readingOrder=2)
GLOBAL_ALIGNMENT_CENTER = Alignment(horizontal='center', vertical='center', readingOrder=2)
GLOBAL_ALIGNMENT_RIGHT_WRAP = Alignment(horizontal='right', vertical='center', wrap_text=True, readingOrder=2)
GLOBAL_ALIGNMENT_CENTER_WRAP = Alignment(horizontal='center', vertical='center', wrap_text=True, readingOrder=2)

FONT_DATA = Font(name='Calibri', size=10, bold=False)
FONT_DATA_BOLD = Font(name='Calibri', size=10, bold=True)
FONT_POS = Font(name='Calibri', size=10, bold=True, color=COLOR_POSITIVE_FONT)
FONT_NEG = Font(name='Calibri', size=10, bold=True, color=COLOR_NEGATIVE_FONT)
FONT_ZERO = Font(name='Calibri', size=10, bold=True, color=COLOR_ZERO_FONT)

FILL_ODD = PatternFill(fill_type='solid', fgColor=COLOR_ROW_ODD)
FILL_EVEN = PatternFill(fill_type='solid', fgColor=COLOR_ROW_EVEN)
FILL_POS = PatternFill(fill_type='solid', fgColor=COLOR_POSITIVE)
FILL_NEG = PatternFill(fill_type='solid', fgColor=COLOR_NEGATIVE)
FILL_ZERO = PatternFill(fill_type='solid', fgColor=COLOR_ZERO)


def _make_border(color=COLOR_BORDER):
    """ایجاد یک شیء Border (فقط برای هدرها یا بخش‌های خاص استفاده شود)"""
    side = Side(style='thin', color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _rtl_alignment(horizontal='right', vertical='center', wrap=False):
    """تنظیم راست‌چین RTL برای سلول‌های فارسی (فقط هدرها)"""
    return Alignment(
        horizontal=horizontal,
        vertical=vertical,
        wrap_text=wrap,
        readingOrder=2,  # 2 = RTL
    )


def _apply_header_style(cell, text, font_size=11, bold=True, bg_color=COLOR_HEADER_BG, font_color=COLOR_HEADER_FONT):
    """اعمال استایل هدر به یک سلول."""
    cell.value = text
    cell.font = Font(name='Calibri', bold=bold, size=font_size, color=font_color)
    cell.fill = PatternFill(fill_type='solid', fgColor=bg_color)
    cell.alignment = _rtl_alignment(horizontal='center')
    cell.border = _make_border(COLOR_BORDER)


def _register_named_styles(wb):
    """ثبت استایل‌های کش شده برای پرهیز از هشینگ مجدد در openpyxl"""
    if 'data_even_right_norm_str' in wb.named_styles:
        return  # قبلا ثبت شده
        
    styles_to_add = []
    
    # 1. Data Styles
    for row_type in ('odd', 'even'):
        for align in ('right', 'center'):
            for bold in (True, False):
                for num_key, num_fmt in [('str', None), ('money', '#,##0.00'), ('pct', '0.00')]:
                    name = f"data_{row_type}_{align}_{'bold' if bold else 'norm'}_{num_key}"
                    ns = NamedStyle(name=name)
                    ns.font = FONT_DATA_BOLD if bold else FONT_DATA
                    ns.fill = FILL_ODD if row_type == 'odd' else FILL_EVEN
                    ns.alignment = GLOBAL_ALIGNMENT_CENTER if align == 'center' else GLOBAL_ALIGNMENT_RIGHT
                    ns.border = GLOBAL_BORDER
                    if num_fmt:
                        ns.number_format = num_fmt
                    styles_to_add.append(ns)

    # 2. Balance Styles
    for bal in ('pos', 'neg', 'zero', 'review'):
        name = f"bal_{bal}"
        ns = NamedStyle(name=name)
        if bal == 'pos':
            ns.font = FONT_POS; ns.fill = FILL_POS
        elif bal == 'neg':
            ns.font = FONT_NEG; ns.fill = FILL_NEG
        elif bal == 'zero':
            ns.font = FONT_ZERO; ns.fill = FILL_ZERO
        else: # review
            ns.font = Font(name='Calibri', size=10, bold=True, color="595959")
            ns.fill = PatternFill(fill_type='solid', fgColor="F2F2F2")
        ns.alignment = GLOBAL_ALIGNMENT_CENTER
        ns.border = GLOBAL_BORDER
        if bal != 'review':
            ns.number_format = '#,##0.00'
        styles_to_add.append(ns)
        
    # 3. Transaction Type Styles
    for txn_type in ('IN', 'OUT'):
        name = f"data_{txn_type.lower()}"
        ns = NamedStyle(name=name)
        if txn_type == 'IN':
            ns.font = FONT_POS; ns.fill = FILL_POS
        else:
            ns.font = FONT_NEG; ns.fill = FILL_NEG
        ns.alignment = GLOBAL_ALIGNMENT_CENTER
        ns.border = GLOBAL_BORDER
        styles_to_add.append(ns)
        
    for s in styles_to_add:
        wb.add_named_style(s)


def _apply_data_style(cell, value, row_idx, number_format=None, bold=False, wrap=False, horizontal='right'):
    """اعمال استایل داده به یک سلول معمولی (نسخه NamedStyle بسیار سریع)."""
    cell.value = value
    
    row_type = 'odd' if row_idx % 2 == 0 else 'even'
    is_bold = 'bold' if bold else 'norm'
    
    if number_format == '#,##0.00':
        num_key = 'money'
    elif number_format == '0.00':
        num_key = 'pct'
    elif number_format:
        num_key = 'str'  # Fallback
    else:
        num_key = 'str'
        
    style_name = f"data_{row_type}_{horizontal}_{is_bold}_{num_key}"
    cell.style = style_name
    
    # Override number_format if it is a custom one not in our standard map
    if number_format and num_key == 'str':
        cell.number_format = number_format
        
    # Wrap text override if needed
    if wrap:
        cell.alignment = GLOBAL_ALIGNMENT_CENTER_WRAP if horizontal == 'center' else GLOBAL_ALIGNMENT_RIGHT_WRAP


def _apply_balance_style(cell, value):
    """اعمال استایل شرطی به ستون موازنه (نسخه NamedStyle بسیار سریع)"""
    if isinstance(value, str) and "در دست بررسی" in value:
        cell.value = value
        cell.style = 'bal_review'
        return
        
    cell.value = float(value)
    if value > 0:
        cell.style = 'bal_pos'
    elif value < 0:
        cell.style = 'bal_neg'
    else:
        cell.style = 'bal_zero'


def _balance_label(value):
    """تبدیل عدد موازنه به برچسب توصیفی فارسی."""
    if isinstance(value, str) and "در دست بررسی" in value:
        return "در دست بررسی ⏳"
    if value > 0:
        return "مازاد پرداخت ✔"
    elif value < 0:
        return "کسری متریال ✘"
    else:
        return "ایده‌آل ✔"


def _write_info_cell(ws, row, col_start, col_end, text, is_label=False):
    """
    نوشتن و استایل‌دهی یک سلول اطلاعاتی در مشخصات بالا.
    """
    if col_start == col_end:
        cell = ws.cell(row=row, column=col_start)
    else:
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
        cell = ws.cell(row=row, column=col_start)
    
    cell.value = text
    cell.alignment = _rtl_alignment(horizontal='right')
    
    if is_label:
        cell.font = Font(name='Calibri', bold=True, size=10, color="1F3864")
        cell.fill = PatternFill(fill_type='solid', fgColor=COLOR_INFO_LABEL_BG)
    else:
        cell.font = Font(name='Calibri', size=10, color="333333")
        cell.fill = PatternFill(fill_type='solid', fgColor=COLOR_INFO_BG)
        
    cell.border = _make_border(COLOR_INFO_BORDER)
    
    # برای سلول‌های merge شده باید حاشیه و بک‌گراند را به کل ستون‌ها اعمال کنیم تا زشت نشود
    for r in range(row, row + 1):
        for c in range(col_start, col_end + 1):
            cell_in_merge = ws.cell(row=r, column=c)
            cell_in_merge.border = _make_border(COLOR_INFO_BORDER)
            if not is_label:
                cell_in_merge.fill = PatternFill(fill_type='solid', fgColor=COLOR_INFO_BG)
            else:
                cell_in_merge.fill = PatternFill(fill_type='solid', fgColor=COLOR_INFO_LABEL_BG)


def _build_contractor_sheet(ws, contractor, rows):
    """
    ساخت یک شیت کامل برای یک پیمانکار.
    """
    ws.sheet_view.rightToLeft = True  # RTL برای کل شیت
    TOTAL_COLS = 15

    # ── ۱. ردیف عنوان اصلی ─────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
    title_cell = ws.cell(row=1, column=1)
    _apply_header_style(
        title_cell,
        "گزارش موازنه متریال - شرکت جهانپارس",
        font_size=16,
        bg_color=COLOR_HEADER_BG,
    )
    ws.row_dimensions[1].height = 38

    # ── ۲. ردیف توضیح فرمول ────────────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=TOTAL_COLS)
    sub_cell = ws.cell(row=2, column=1)
    sub_cell.value = "موازنه = کل تحویلی − (کار تاییدشده + پرتی مجاز)    |    سبز: مازاد پرداخت    |    قرمز: کسری متریال    |    زرد: ایده‌آل"
    sub_cell.font  = Font(name='Calibri', size=9, italic=True, color="4472C4")
    sub_cell.fill  = PatternFill(fill_type='solid', fgColor="EBF3FB")
    sub_cell.alignment = _rtl_alignment(horizontal='center')
    ws.row_dimensions[2].height = 18

    # ── ۳. بلوک اطلاعات پیمانکار (ردیف ۳) ────────────────────────
    contractor_name = contractor.get_full_name()

    # ردیف ۳: نام پیمانکار، تاریخ گزارش
    _write_info_cell(ws, 3, 1, 1, "نام پیمانکار :", is_label=True)
    _write_info_cell(ws, 3, 2, 8, contractor_name, is_label=False)
    
    _write_info_cell(ws, 3, 9, 10, "تاریخ گزارش :", is_label=True)
    _write_info_cell(ws, 3, 11, 15, str(jdatetime.date.today()), is_label=False)
    
    ws.row_dimensions[3].height = 20

    # ── ۴. خط جداکننده ردیف ۴ ──────────────────────────────────────────────
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=TOTAL_COLS)
    sep = ws.cell(row=4, column=1)
    sep.fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
    ws.row_dimensions[4].height = 6

    # ── ۵. هدر ستون‌های جدول (ردیف ۵) ─────────────────────────────────────
    HEADERS = [
        "ردیف",
        "شماره قرارداد",
        "موضوع قرارداد",
        "رسته کاری",
        "نام کالا",
        "سایز",
        "جنس",
        "ضخامت",
        "واحد",
        "کل متریال\nتحویلی",
        "مقدار کار\nتاییدشده",
        "درصد\nپرتی (%)",
        "پرتی\nمجاز",
        "موازنه\n(انحراف)",
        "وضعیت نهایی",
    ]
    COL_WIDTHS = [6, 16, 25, 16, 28, 12, 14, 12, 8, 14, 14, 10, 12, 14, 20]

    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=5, column=col_idx)
        _apply_header_style(cell, header, font_size=10, bg_color=COLOR_SUB_HEADER_BG)
        cell.alignment = _rtl_alignment(horizontal='center', wrap=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[5].height = 38
    ws.freeze_panes = "A6"

    # ── ۶. ردیف‌های داده (از ردیف ۶) ─────────────────────────────────────
    NUM_FMT = '#,##0.00'

    for row_offset, row_data in enumerate(rows):
        excel_row = row_offset + 6

        _apply_data_style(ws.cell(excel_row, 1),  row_offset + 1,                    row_offset, horizontal='center')
        _apply_data_style(ws.cell(excel_row, 2),  row_data['contract_number'],        row_offset, horizontal='center')
        _apply_data_style(ws.cell(excel_row, 3),  row_data['contract_subject'],       row_offset)
        _apply_data_style(ws.cell(excel_row, 4),  row_data['work_category'],          row_offset)
        _apply_data_style(ws.cell(excel_row, 5),  row_data['material_name'],          row_offset)
        _apply_data_style(ws.cell(excel_row, 6),  row_data['size'],                   row_offset, horizontal='center')
        _apply_data_style(ws.cell(excel_row, 7),  row_data['mat_type'],               row_offset, horizontal='center')
        _apply_data_style(ws.cell(excel_row, 8),  row_data['thickness'],              row_offset, horizontal='center')
        _apply_data_style(ws.cell(excel_row, 9),  row_data['unit'],                   row_offset, horizontal='center')
        _apply_data_style(ws.cell(excel_row, 10), float(row_data['total_issued']),    row_offset, number_format=NUM_FMT)
        _apply_data_style(ws.cell(excel_row, 11), float(row_data['approved_work']),   row_offset, number_format=NUM_FMT)
        _apply_data_style(ws.cell(excel_row, 12), row_data['waste_pct'],              row_offset, number_format='0.00')
        _apply_data_style(ws.cell(excel_row, 13), float(row_data['allowed_waste']),   row_offset, number_format=NUM_FMT)
        _apply_balance_style(ws.cell(excel_row, 14), row_data['balance'])

        # وضعیت نهایی
        status_cell = ws.cell(excel_row, 15)
        status_cell.value = row_data['balance_label']
        if isinstance(row_data['balance'], str):
            status_cell.style = 'bal_review'
        elif row_data['balance'] > 0:
            status_cell.style = 'bal_pos'
        elif row_data['balance'] < 0:
            status_cell.style = 'bal_neg'
        else:
            status_cell.style = 'bal_zero'
            
        ws.row_dimensions[excel_row].height = 22

    # ── ۷. ردیف جمع کل ─────────────────────────────────────────────────────
    if rows:
        total_row = len(rows) + 6
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=9)
        _apply_header_style(ws.cell(total_row, 1), "جمع کل", font_size=11, bg_color=COLOR_HEADER_BG)

        grand_issued   = sum(r['total_issued']  for r in rows)
        grand_approved = sum(r['approved_work'] for r in rows)
        grand_waste    = sum(r['allowed_waste'] for r in rows)
        grand_balance  = sum(r['balance']       for r in rows if not isinstance(r['balance'], str))

        for col, val in [(10, grand_issued), (11, grand_approved), (13, grand_waste)]:
            c = ws.cell(total_row, col)
            _apply_header_style(c, float(val), bg_color=COLOR_HEADER_BG)
            c.number_format = NUM_FMT

        # ستون درصد پرتی در ردیف جمع خالی ولی با رنگ
        ws.cell(total_row, 12).fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
        ws.cell(total_row, 12).border = GLOBAL_BORDER

        _apply_balance_style(ws.cell(total_row, 14), grand_balance)
        ws.cell(total_row, 14).fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
        ws.cell(total_row, 14).font = Font(name='Calibri', size=11, bold=True, color=COLOR_HEADER_FONT)

        ws.cell(total_row, 15).fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
        ws.cell(total_row, 15).border = GLOBAL_BORDER
        ws.row_dimensions[total_row].height = 26

        # ── ۸. AutoFilter روی جدول داده ─────────────────────────────────────
        last_data_row = total_row - 1
        first_col_letter = get_column_letter(1)
        last_col_letter = get_column_letter(TOTAL_COLS)
        ws.auto_filter.ref = f"{first_col_letter}5:{last_col_letter}{last_data_row}"

    # ── ۹. فریز پنل (هدر و اطلاعات پیمانکار ثابت بماند) ────────────────────
    ws.freeze_panes = "A6"
    ws.print_title_rows = '1:5'
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1


def _build_global_sheet(ws, rows, wb, task_id=None, start_row_idx=0, phase_start_time=None):
    """
    ساخت شیت گزارش موازنه کل کارگاه (بهینه‌شده با xlsxwriter)
    """
    COLOR_HEADER_BG = "#1F3864"
    COLOR_HEADER_FONT = "#FFFFFF"
    COLOR_INFO_BG = "#EBF3FB"
    COLOR_BORDER = "#BDD7EE"
    
    header_fmt = wb.add_format({
        'font_name': 'Calibri', 'font_size': 11, 'bold': True,
        'bg_color': COLOR_HEADER_BG, 'font_color': COLOR_HEADER_FONT,
        'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': COLOR_BORDER, 'reading_order': 2, 'text_wrap': True
    })
    
    title_fmt = wb.add_format({
        'font_name': 'Calibri', 'font_size': 16, 'bold': True,
        'align': 'center', 'valign': 'vcenter', 'reading_order': 2
    })
    
    info_fmt = wb.add_format({
        'font_name': 'Calibri', 'font_size': 9, 'italic': True, 'font_color': '#4472C4',
        'bg_color': COLOR_INFO_BG, 'align': 'center', 'valign': 'vcenter', 'reading_order': 2
    })
    
    label_fmt = wb.add_format({
        'font_name': 'Calibri', 'font_size': 10, 'bold': True,
        'bg_color': '#D6E4F0', 'align': 'right', 'valign': 'vcenter', 'border': 1, 'border_color': COLOR_BORDER, 'reading_order': 2
    })
    
    val_fmt = wb.add_format({
        'font_name': 'Calibri', 'font_size': 10,
        'bg_color': COLOR_INFO_BG, 'align': 'right', 'valign': 'vcenter', 'border': 1, 'border_color': COLOR_BORDER, 'reading_order': 2
    })
    
    data_fmt_center = wb.add_format({'font_name': 'Calibri', 'font_size': 10, 'align': 'center', 'valign': 'vcenter', 'reading_order': 2})
    data_fmt_right = wb.add_format({'font_name': 'Calibri', 'font_size': 10, 'align': 'right', 'valign': 'vcenter', 'reading_order': 2})
    data_fmt_num = wb.add_format({'font_name': 'Calibri', 'font_size': 10, 'align': 'center', 'valign': 'vcenter', 'num_format': '#,##0.00', 'reading_order': 2})
    
    bal_pos_fmt = wb.add_format({'font_name': 'Calibri', 'font_size': 10, 'bold': True, 'font_color': '#276221', 'bg_color': '#C6EFCE', 'align': 'center', 'valign': 'vcenter', 'num_format': '#,##0.00', 'reading_order': 2})
    bal_neg_fmt = wb.add_format({'font_name': 'Calibri', 'font_size': 10, 'bold': True, 'font_color': '#9C0006', 'bg_color': '#FFC7CE', 'align': 'center', 'valign': 'vcenter', 'num_format': '#,##0.00', 'reading_order': 2})
    bal_zero_fmt = wb.add_format({'font_name': 'Calibri', 'font_size': 10, 'bold': True, 'font_color': '#9C6500', 'bg_color': '#FFEB9C', 'align': 'center', 'valign': 'vcenter', 'num_format': '#,##0.00', 'reading_order': 2})
    
    sum_fmt_val = wb.add_format({'font_name': 'Calibri', 'font_size': 11, 'bold': True, 'bg_color': COLOR_HEADER_BG, 'font_color': COLOR_HEADER_FONT, 'align': 'center', 'valign': 'vcenter', 'num_format': '#,##0.00', 'border': 1, 'border_color': COLOR_BORDER, 'reading_order': 2})
    sum_fmt_lbl = wb.add_format({'font_name': 'Calibri', 'font_size': 11, 'bold': True, 'bg_color': COLOR_HEADER_BG, 'font_color': COLOR_HEADER_FONT, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'border_color': COLOR_BORDER, 'reading_order': 2})

    ws.right_to_left()
    
    if start_row_idx == 0:
        ws.merge_range('A1:P1', "گزارش موازنه متریال کل - شرکت جهانپارس", title_fmt)
        ws.merge_range('A2:P2', "موازنه = کل تحویلی − (کار تاییدشده + پرتی مجاز) | سبز: مازاد | قرمز: کسری | زرد: ایده‌آل", info_fmt)
        ws.write(2, 0, "تاریخ گزارش :", label_fmt)
        ws.write(2, 1, str(jdatetime.date.today()), val_fmt)
        ws.write(3, 0, "حوزه گزارش :", label_fmt)
        ws.write(3, 1, "موازنه کل (به تفکیک تمامی پیمانکاران)", val_fmt)

        for col in range(2, 16):
            ws.write(2, col, "", wb.add_format({'bg_color': COLOR_HEADER_BG}))
            ws.write(3, col, "", wb.add_format({'bg_color': COLOR_HEADER_BG}))
            ws.write(4, col, "", wb.add_format({'bg_color': COLOR_HEADER_BG}))
        ws.write(4, 0, "", wb.add_format({'bg_color': COLOR_HEADER_BG}))
        ws.write(4, 1, "", wb.add_format({'bg_color': COLOR_HEADER_BG}))

        HEADERS = [
            "ردیف", "پیمانکار", "شماره قرارداد", "موضوع قرارداد", "رسته کاری", "نام کالا", 
            "سایز", "جنس", "ضخامت", "واحد", "کل متریال\nتحویلی", "مقدار کار\nتاییدشده", 
            "درصد\nپرتی (%)", "پرتی\nمجاز", "موازنه\n(انحراف)", "وضعیت نهایی"
        ]
        ws.set_row(5, 30)
        ws.write_row(5, 0, HEADERS, header_fmt)
        
    import time
    total_rows = len(rows)
    current_row = 6 if start_row_idx == 0 else start_row_idx
    
    for row_offset in range(start_row_idx, total_rows):
        row_data = rows[row_offset]
        if task_id and total_rows > 0 and (row_offset % 2000 == 0 or row_offset == total_rows - 1):
            ratio = (row_offset + 1) / total_rows
            progress_pct = 20 + int(ratio * 75)  # maps 0→1 to 20→95

            # ETA calculation
            eta_seconds = None
            if phase_start_time and ratio > 0.01:
                elapsed = time.time() - phase_start_time
                remaining = elapsed * (1 - ratio) / ratio
                eta_seconds = max(1, int(remaining))

            # Cancel check (every 5000 rows only)
            if row_offset % 5000 == 0:
                from .models import ExportTask
                task_obj = ExportTask.objects.filter(pk=task_id).first()
                if task_obj and task_obj.status not in ('PENDING', 'PROCESSING'):
                    raise ValueError("توسط کاربر لغو شد.")

            ExportTask.objects.filter(pk=task_id).update(progress=progress_pct, eta=eta_seconds or 0)
            from .tasks import send_task_progress
            send_task_progress(task_id, progress_pct, phase='در حال ساخت فایل اکسل...', eta=eta_seconds)
        
        balance = row_data['balance']
        if isinstance(balance, (int, float)):
            if balance > 0:
                bal_fmt = bal_pos_fmt
                status_fmt = bal_pos_fmt
            elif balance < 0:
                bal_fmt = bal_neg_fmt
                status_fmt = bal_neg_fmt
            else:
                bal_fmt = bal_zero_fmt
                status_fmt = bal_zero_fmt
        else:
            bal_fmt = data_fmt_center
            status_fmt = data_fmt_center

        ws.write(current_row, 0, row_offset + 1, data_fmt_center)
        ws.write(current_row, 1, row_data['contractor_name'], data_fmt_right)
        ws.write(current_row, 2, row_data['contract_number'], data_fmt_center)
        ws.write(current_row, 3, row_data['contract_subject'], data_fmt_right)
        ws.write(current_row, 4, row_data['work_category'], data_fmt_right)
        ws.write(current_row, 5, row_data['material_name'], data_fmt_right)
        ws.write(current_row, 6, row_data['size'], data_fmt_center)
        ws.write(current_row, 7, row_data['mat_type'], data_fmt_center)
        ws.write(current_row, 8, row_data['thickness'], data_fmt_center)
        ws.write(current_row, 9, row_data['unit'], data_fmt_center)
        ws.write(current_row, 10, float(row_data['total_issued']), data_fmt_num)
        ws.write(current_row, 11, float(row_data['approved_work']), data_fmt_num)
        ws.write(current_row, 12, float(row_data['waste_pct']), data_fmt_num)
        ws.write(current_row, 13, float(row_data['allowed_waste']), data_fmt_num)
        ws.write(current_row, 14, balance, bal_fmt)
        ws.write(current_row, 15, row_data['balance_label'], status_fmt)
        
        current_row += 1

    if rows:
        grand_issued   = sum(r['total_issued']  for r in rows)
        grand_approved = sum(r['approved_work'] for r in rows)
        grand_waste    = sum(r['allowed_waste'] for r in rows)
        grand_balance  = sum(r['balance']       for r in rows if not isinstance(r['balance'], str))
        
        for c in range(16):
            ws.write(current_row, c, "", sum_fmt_lbl)
        
        ws.write(current_row, 0, "جمع کل", sum_fmt_lbl)
        ws.write(current_row, 10, float(grand_issued), sum_fmt_val)
        ws.write(current_row, 11, float(grand_approved), sum_fmt_val)
        ws.write(current_row, 13, float(grand_waste), sum_fmt_val)
        
        if grand_balance > 0:
            bal_sum_fmt = wb.add_format({'font_name': 'Calibri', 'font_size': 11, 'bold': True, 'font_color': '#276221', 'bg_color': COLOR_HEADER_BG, 'align': 'center', 'valign': 'vcenter', 'num_format': '#,##0.00', 'border': 1, 'border_color': COLOR_BORDER})
        elif grand_balance < 0:
            bal_sum_fmt = wb.add_format({'font_name': 'Calibri', 'font_size': 11, 'bold': True, 'font_color': '#9C0006', 'bg_color': COLOR_HEADER_BG, 'align': 'center', 'valign': 'vcenter', 'num_format': '#,##0.00', 'border': 1, 'border_color': COLOR_BORDER})
        else:
            bal_sum_fmt = sum_fmt_val
            
        ws.write(current_row, 14, float(grand_balance), bal_sum_fmt)


# ─────────────────────────────────────────────────────────────────────────────
# تابع اصلی تولید گزارش
# ─────────────────────────────────────────────────────────────────────────────
def generate_material_balance_excel(
    contractor_id: int | None = None,
    material_id: int | None = None,
    is_superuser: bool = False,
) -> bytes:
    """
    تولید گزارش موازنه متریال در قالب فایل اکسل.

    آرگومان‌های اختیاری:
        contractor_id: در صورت ارائه، فقط برای آن پیمانکار فیلتر می‌شود.
        material_id:   در صورت ارائه، فقط برای آن متریال فیلتر می‌شود.

    خروجی:
        bytes - محتوای فایل xlsx آماده دانلود یا ذخیره.
        یک شیت جداگانه به ازای هر پیمانکار.

    فرمول‌های محاسباتی:
        ۱. کل متریال تحویلی  = جمع OUT از WarehouseTransaction
        ۲. مقدار کار تاییدشده = جمع approved_quantity از TechnicalOfficeApproval
        ۳. پرتی مجاز         = مقدار کار تاییدشده × (waste_percentage / 100)
        ۴. موازنه            = کل تحویلی - (تاییدشده + پرتی مجاز)
    """
    # import داخلی برای جلوگیری از مشکلات circular import در Django
    from .models import WarehouseTransaction, TechnicalOfficeApproval, MaterialItem, Contractor

    # ─── ۱. بارگذاری و جمع‌بندی داده‌های تراکنش خروجی انبار ───────────────────────────
    issue_qs = WarehouseTransaction.objects.filter(transaction_type='OUT')
    if contractor_id:
        issue_qs = issue_qs.filter(contractor_id=contractor_id)
    if material_id:
        issue_qs = issue_qs.filter(material_id=material_id)

    # جمع‌بندی مقادیر با دیتابیس (سرعت هزار برابری)
    issued_aggs = issue_qs.values('contractor_id', 'material_id', 'contract_number', 'contract_subject').annotate(total_qty=Sum('quantity'))
    issued_map: dict[tuple, Decimal] = {}
    for agg in issued_aggs:
        k = (agg['contractor_id'], agg['material_id'], agg['contract_number'] or '', agg['contract_subject'] or '')
        issued_map[k] = agg['total_qty'] or Decimal('0')

    # ─── ۲. بارگذاری تاییدیه‌های دفتر فنی ─────────────────────────────────
    approval_qs = TechnicalOfficeApproval.objects.all()
    if contractor_id:
        approval_qs = approval_qs.filter(contractor_id=contractor_id)
    if material_id:
        approval_qs = approval_qs.filter(material_id=material_id)

    appr_aggs = approval_qs.values('contractor_id', 'material_id', 'contract_number', 'contract_subject').annotate(total_appr=Sum('approved_quantity'))
    approved_map: dict[tuple, Decimal] = {}
    for agg in appr_aggs:
        k = (agg['contractor_id'], agg['material_id'], agg['contract_number'] or '', agg['contract_subject'] or '')
        approved_map[k] = agg['total_appr'] or Decimal('0')

    # ─── ۳. جمع‌آوری تمام جفت‌های (contractor, material) یکتا ─────────────
    all_keys = set(issued_map.keys()) | set(approved_map.keys())

    if not all_keys:
        # اگر هیچ داده‌ای وجود نداشت، یک فایل خالی با پیام برگردان
        wb = Workbook()
        ws = wb.active
        ws.title = "بدون داده"
        ws.sheet_view.rightToLeft = True
        ws.merge_cells('A1:E1')
        c = ws['A1']
        c.value = "هیچ داده‌ای برای گزارش‌گیری وجود ندارد."
        c.font = Font(name='Calibri', size=12, bold=True, color="9C0006")
        c.alignment = _rtl_alignment(horizontal='center')
        if not is_superuser:
            ws.protection.sheet = True
            ws.protection.password = "jahanpars2026"
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # بارگذاری آبجکت‌های Contractor و MaterialItem
    contractor_ids = {k[0] for k in all_keys if k[0] is not None}
    material_ids   = {k[1] for k in all_keys}

    contractors = {u.id: u for u in Contractor.objects.filter(id__in=contractor_ids)}
    materials   = {m.id: m for m in MaterialItem.objects.select_related('work_category').filter(id__in=material_ids)}

    # ─── ۴. گروه‌بندی ردیف‌ها بر اساس پیمانکار ─────────────────────────────
    # ساختار: {contractor_id: [row_dict, ...]}
    contractor_rows: dict[int, list] = {c_id: [] for c_id in contractor_ids}

    for (c_id, m_id, contract_num, contract_subj) in sorted(all_keys, key=lambda k: (k[0] or 0, k[1], k[2], k[3])):
        material = materials.get(m_id)
        if not material:
            continue

        total_issued  = issued_map.get((c_id, m_id, contract_num, contract_subj), Decimal('0'))
        
        # بررسی اینکه آیا اصلا تاییدیه‌ای برای این قرارداد و متریال وجود دارد یا خیر
        if (c_id, m_id, contract_num, contract_subj) not in approved_map:
            approved_work = Decimal('0')
            allowed_waste = Decimal('0')
            balance = "در دست بررسی برای تایید دفتر فنی"
        else:
            approved_work = approved_map[(c_id, m_id, contract_num, contract_subj)]
            waste_pct     = material.waste_percentage / Decimal('100')
            allowed_waste = (approved_work * waste_pct).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            balance       = (total_issued - (approved_work + allowed_waste)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

        row = {
            'contract_number': contract_num or "—",
            'contract_subject': contract_subj or "—",
            'work_category':  material.work_category.name if material.work_category else "—",
            'material_name':  material.name,
            'size':           material.size or "—",
            'mat_type':       material.material_type or "—",
            'thickness':      material.thickness or "—",
            'unit':           material.get_unit_display(),
            'total_issued':   total_issued,
            'approved_work':  approved_work,
            'waste_pct':      float(material.waste_percentage),
            'allowed_waste':  allowed_waste,
            'balance':        balance,
            'balance_label':  _balance_label(balance),
        }
        if c_id in contractor_rows:
            contractor_rows[c_id].append(row)

    # ─── ۵. ساخت فایل اکسل - یک شیت به ازای هر پیمانکار ───────────────────
    wb = Workbook()
    _register_named_styles(wb)
    wb.remove(wb.active)  # حذف شیت پیش‌فرض خالی

    for c_id in sorted(contractor_rows.keys()):
        contractor = contractors.get(c_id)
        if not contractor:
            continue

        rows = contractor_rows[c_id]
        sheet_name = contractor.get_full_name()[:31]  # حداکثر ۳۱ کاراکتر
        ws = wb.create_sheet(title=sheet_name)

        _build_contractor_sheet(ws, contractor, rows)
        
        if not is_superuser:
            ws.protection.sheet = True
            ws.protection.password = "jahanpars2026"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def get_global_material_balance_rows_data(
    search=None,
    category=None,
    contractor=None,
    material=None,
    status=None,
    page=1,
    page_size=None,
    return_filters=False
) -> list[dict] | dict:
    """
    دریافت اطلاعات موازنه متریال کل کارگاه به صورت تجمیعی با پشتیبانی از فیلتر، جستجو و صفحه‌بندی (نسخه بهینه‌شده با جدول پیش‌محاسبه).
    """
    from .models import GlobalMaterialBalance, WorkCategory, Contractor, MaterialItem
    from django.db.models import Q

    # 1. Base Queryset with select_related for performance
    qs = GlobalMaterialBalance.objects.select_related('contractor', 'material', 'material__work_category').all()

    # 2. Apply Filters (contractor, material, category, status, search)
    if contractor:
        parts = contractor.split()
        if len(parts) >= 2:
            qs = qs.filter(contractor__first_name__icontains=parts[0], contractor__last_name__icontains=parts[1])
        else:
            qs = qs.filter(Q(contractor__first_name__icontains=contractor) | Q(contractor__last_name__icontains=contractor))

    if material:
        qs = qs.filter(material__name__icontains=material)

    if category:
        qs = qs.filter(material__work_category__name__icontains=category)

    if status:
        # Map frontend status filter strings to stored database balance_label values
        status_map = {
            "کسری متریال (بدهکار به کارفرما)": "کسری متریال ✘",
            "مازاد پرداخت (جنس اضافه نزد پیمانکار)": "مازاد پرداخت ✔",
            "موازنه ایده‌آل (بدون انحراف)": "ایده‌آل ✔",
            "در دست بررسی برای تایید دفتر فنی": "در دست بررسی ⏳"
        }
        mapped_status = status_map.get(status, status)
        qs = qs.filter(balance_label=mapped_status)

    if search:
        search_q = Q(material__name__icontains=search) | \
                   Q(contractor__first_name__icontains=search) | \
                   Q(contractor__last_name__icontains=search) | \
                   Q(contract_number__icontains=search) | \
                   Q(contract_subject__icontains=search)
        qs = qs.filter(search_q)

    # 3. Sort (Same sorting order as previous version: contractor, material, contract_number, contract_subject)
    qs = qs.order_by(
        'contractor__first_name', 'contractor__last_name',
        'material__name',
        'contract_number',
        'contract_subject'
    )

    # 4. Build rows and Paginate
    if page_size is None:
        # Excel/CSV mode: use .values().iterator() to prevent loading all model objects at once
        qs_values = qs.values(
            'contractor__first_name',
            'contractor__last_name',
            'material__name',
            'material__size',
            'material__material_type',
            'material__thickness',
            'material__unit',
            'material__waste_percentage',
            'material__work_category__name',
            'contract_number',
            'contract_subject',
            'total_issued',
            'approved_work',
            'allowed_waste',
            'balance',
            'balance_label'
        ).iterator(chunk_size=5000)
        UNIT_MAP = {
            'KG': 'کیلوگرم',
            'M': 'متر',
            'SQM': 'متر مربع',
            'PCS': 'عدد',
        }
        rows = []
        for row in qs_values:
            balance_val = row['balance']
            first_name = row['contractor__first_name'] or ''
            last_name = row['contractor__last_name'] or ''
            full_name = f"{first_name} {last_name}".strip()
            rows.append({
                'contractor_name': full_name if full_name else "—",
                'contract_number': row['contract_number'] or "—",
                'contract_subject': row['contract_subject'] or "—",
                'work_category':  row['material__work_category__name'] or "—",
                'material_name':  row['material__name'],
                'size':           row['material__size'] or "—",
                'mat_type':       row['material__material_type'] or "—",
                'thickness':      row['material__thickness'] or "—",
                'unit':           UNIT_MAP.get(row['material__unit'], row['material__unit'] or "—"),
                'total_issued':   float(row['total_issued']),
                'approved_work':  float(row['approved_work']),
                'waste_pct':      float(row['material__waste_percentage'] or 0),
                'allowed_waste':  float(row['allowed_waste']),
                'balance':        float(balance_val) if balance_val is not None else "در دست بررسی برای تایید دفتر فنی",
                'balance_label':  row['balance_label'],
            })
        return rows
    else:
        # Pagination mode: slice query and load only 10 model objects
        total_count = qs.count()
        start = (page - 1) * page_size
        end = start + page_size
        page_qs = qs[start:end]

        rows = []
        for row in page_qs:
            mat = row.material
            cont = row.contractor
            work_category = mat.work_category

            rows.append({
                'contractor_name': cont.get_full_name() if cont else "—",
                'contract_number': row.contract_number or "—",
                'contract_subject': row.contract_subject or "—",
                'work_category':  work_category.name if work_category else "—",
                'material_name':  mat.name,
                'size':           mat.size or "—",
                'mat_type':       mat.material_type or "—",
                'thickness':      mat.thickness or "—",
                'unit':           mat.get_unit_display(),
                'total_issued':   float(row.total_issued),
                'approved_work':  float(row.approved_work),
                'waste_pct':      float(mat.waste_percentage),
                'allowed_waste':  float(row.allowed_waste),
                'balance':        float(row.balance) if row.balance is not None else "در دست بررسی برای تایید دفتر فنی",
                'balance_label':  row.balance_label,
            })

    # Fetch unique filters (only if return_filters=True, typically on initial load)
    if return_filters:
        # Optimizing filter options using distinct values
        cats_list = list(WorkCategory.objects.filter(materials__global_balances__isnull=False).values_list('name', flat=True).distinct().order_by('name'))
        conts_list = [f"{c['first_name']} {c['last_name']}".strip() for c in Contractor.objects.filter(global_balances__isnull=False).values('first_name', 'last_name').distinct().order_by('first_name', 'last_name')]
        mats_list = list(MaterialItem.objects.filter(global_balances__isnull=False).values_list('name', flat=True).distinct().order_by('name'))
    else:
        cats_list, conts_list, mats_list = [], [], []

    return {
        'count': total_count,
        'results': rows,
        'filters': {
            'categories': cats_list,
            'contractors': conts_list,
            'materials': mats_list,
            'statuses': ["کسری متریال (بدهکار به کارفرما)", "مازاد پرداخت (جنس اضافه نزد پیمانکار)", "موازنه ایده‌آل (بدون انحراف)", "در دست بررسی برای تایید دفتر فنی"]
        }
    }


def update_global_balance_for_key(contractor_id, material_id, contract_number, contract_subject):
    """
    محاسبه موازنه برای یک کلید خاص و ذخیره یا به‌روزرسانی آن در جدول پیش‌محاسبه شده.
    """
    from .models import WarehouseTransaction, TechnicalOfficeApproval, MaterialItem, GlobalMaterialBalance
    from django.db.models import Sum
    from decimal import Decimal, ROUND_HALF_UP

    contract_number = contract_number or ''
    contract_subject = contract_subject or ''

    if not contractor_id or not material_id:
        return

    # 1. محاسبه جمع صادر شده (تراکنش خروج)
    total_issued = WarehouseTransaction.objects.filter(
        transaction_type='OUT',
        contractor_id=contractor_id,
        material_id=material_id,
        contract_number=contract_number,
        contract_subject=contract_subject
    ).aggregate(total=Sum('quantity'))['total'] or Decimal('0')

    # 2. بررسی تاییدیه فنی
    approval_qs = TechnicalOfficeApproval.objects.filter(
        contractor_id=contractor_id,
        material_id=material_id,
        contract_number=contract_number,
        contract_subject=contract_subject
    )

    is_approved = approval_qs.exists()

    # اگر تراکنشی وجود نداشت و تاییدیه‌ای هم نبود، رکورد قبلی را پاک می‌کنیم
    if total_issued == Decimal('0') and not is_approved:
        GlobalMaterialBalance.objects.filter(
            contractor_id=contractor_id,
            material_id=material_id,
            contract_number=contract_number,
            contract_subject=contract_subject
        ).delete()
        return

    try:
        material = MaterialItem.objects.get(pk=material_id)
    except MaterialItem.DoesNotExist:
        return

    if not is_approved:
        approved_work = Decimal('0')
        allowed_waste = Decimal('0')
        balance = None
        balance_label = "در دست بررسی ⏳"
    else:
        approved_work = approval_qs.aggregate(total=Sum('approved_quantity'))['total'] or Decimal('0')
        waste_pct = material.waste_percentage / Decimal('100')
        allowed_waste = (approved_work * waste_pct).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        balance = (total_issued - (approved_work + allowed_waste)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        balance_label = _balance_label(balance)

    # به‌روزرسانی یا ایجاد در جدول پیش‌محاسبه
    GlobalMaterialBalance.objects.update_or_create(
        contractor_id=contractor_id,
        material_id=material_id,
        contract_number=contract_number,
        contract_subject=contract_subject,
        defaults={
            'total_issued': total_issued,
            'approved_work': approved_work,
            'allowed_waste': allowed_waste,
            'balance': balance,
            'balance_label': balance_label
        }
    )


def recalculate_all_balances_for_material(material_id):
    """
    به‌روزرسانی موازنه تمام کلیدهای مربوط به یک متریال خاص (زمانی که درصد پرتی کالا تغییر کند)
    با استفاده از bulk_update برای بهینه‌سازی.
    """
    from .models import GlobalMaterialBalance, MaterialItem
    from decimal import Decimal, ROUND_HALF_UP

    try:
        material = MaterialItem.objects.get(pk=material_id)
        waste_pct = material.waste_percentage / Decimal('100')
    except MaterialItem.DoesNotExist:
        return

    balances = GlobalMaterialBalance.objects.filter(material_id=material_id)
    to_update = []

    for b in balances:
        allowed_waste = (b.approved_work * waste_pct).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        balance_val = (b.total_issued - (b.approved_work + allowed_waste)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        
        b.allowed_waste = allowed_waste
        b.balance = balance_val
        b.balance_label = _balance_label(balance_val)
        to_update.append(b)

    if to_update:
        GlobalMaterialBalance.objects.bulk_update(
            to_update,
            ['allowed_waste', 'balance', 'balance_label'],
            batch_size=1000
        )


def rebuild_all_global_balances():
    """
    پاک کردن و بازسازی کامل جدول پیش‌محاسبه شده موازنه کل از روی داده‌های خام تراکنش‌ها و تاییدیه‌ها.
    """
    from .models import WarehouseTransaction, TechnicalOfficeApproval, MaterialItem, GlobalMaterialBalance
    from django.db import transaction
    from django.db.models import Sum
    from decimal import Decimal, ROUND_HALF_UP

    # 1. دریافت تمام داده‌های خام تجمیع شده
    issue_qs = WarehouseTransaction.objects.filter(transaction_type='OUT')
    approval_qs = TechnicalOfficeApproval.objects.all()

    issued_aggs = list(issue_qs.values('contractor_id', 'material_id', 'contract_number', 'contract_subject').annotate(total_qty=Sum('quantity')))
    appr_aggs = list(approval_qs.values('contractor_id', 'material_id', 'contract_number', 'contract_subject').annotate(total_appr=Sum('approved_quantity')))

    issued_map = {(agg['contractor_id'], agg['material_id'], agg['contract_number'] or '', agg['contract_subject'] or ''): agg['total_qty'] or Decimal('0') for agg in issued_aggs}
    approved_map = {(agg['contractor_id'], agg['material_id'], agg['contract_number'] or '', agg['contract_subject'] or ''): agg['total_appr'] or Decimal('0') for agg in appr_aggs}

    all_keys = set(issued_map.keys()) | set(approved_map.keys())

    material_ids = {k[1] for k in all_keys}
    materials_waste = {m['id']: m['waste_percentage'] for m in MaterialItem.objects.filter(id__in=material_ids).values('id', 'waste_percentage')}

    precomputed_instances = []
    for c_id, m_id, contract_num, contract_subj in all_keys:
        if not c_id or not m_id:
            continue

        total_issued = issued_map.get((c_id, m_id, contract_num, contract_subj), Decimal('0'))

        if (c_id, m_id, contract_num, contract_subj) not in approved_map:
            approved_work = Decimal('0')
            allowed_waste = Decimal('0')
            balance = None
            balance_label = "در دست بررسی ⏳"
        else:
            approved_work = approved_map[(c_id, m_id, contract_num, contract_subj)]
            waste_pct = materials_waste.get(m_id, Decimal('0')) / Decimal('100')
            allowed_waste = (approved_work * waste_pct).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            balance = (total_issued - (approved_work + allowed_waste)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            balance_label = _balance_label(balance)

        precomputed_instances.append(GlobalMaterialBalance(
            contractor_id=c_id,
            material_id=m_id,
            contract_number=contract_num,
            contract_subject=contract_subj,
            total_issued=total_issued,
            approved_work=approved_work,
            allowed_waste=allowed_waste,
            balance=balance,
            balance_label=balance_label
        ))

    with transaction.atomic():
        GlobalMaterialBalance.objects.all().delete()
        GlobalMaterialBalance.objects.bulk_create(precomputed_instances, batch_size=5000)


# ─────────────────────────────────────────────────────────────────────────────
# توابع کمکی write-only برای شیت موازنه کل (Phase 1: Streaming Excel)
# ─────────────────────────────────────────────────────────────────────────────

def _wo_cell(ws, value, font=None, fill=None, alignment=None, border=None, number_format=None):
    """ساخت یک WriteOnlyCell با استایل مشخص."""
    cell = WriteOnlyCell(ws, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    return cell


def _wo_header_cell(ws, text, font_size=11, bg_color=COLOR_HEADER_BG, font_color=COLOR_HEADER_FONT):
    """ساخت یک WriteOnlyCell با استایل هدر."""
    return _wo_cell(
        ws, text,
        font=Font(name='Calibri', bold=True, size=font_size, color=font_color),
        fill=PatternFill(fill_type='solid', fgColor=bg_color),
        alignment=GLOBAL_ALIGNMENT_CENTER,
        border=GLOBAL_BORDER,
    )


def _wo_info_cell(ws, text, is_label=False):
    """ساخت یک WriteOnlyCell برای بلوک اطلاعات."""
    if is_label:
        return _wo_cell(
            ws, text,
            font=Font(name='Calibri', bold=True, size=10, color="1F3864"),
            fill=PatternFill(fill_type='solid', fgColor=COLOR_INFO_LABEL_BG),
            alignment=GLOBAL_ALIGNMENT_RIGHT,
            border=_make_border(COLOR_INFO_BORDER),
        )
    else:
        return _wo_cell(
            ws, text,
            font=Font(name='Calibri', size=10, color="333333"),
            fill=PatternFill(fill_type='solid', fgColor=COLOR_INFO_BG),
            alignment=GLOBAL_ALIGNMENT_RIGHT,
            border=_make_border(COLOR_INFO_BORDER),
        )


def _wo_data_cell(ws, value, row_idx, number_format=None, horizontal='right'):
    """ساخت یک WriteOnlyCell برای داده‌های معمولی (جایگزین _apply_data_style در write-only mode)."""
    row_type_fill = FILL_ODD if row_idx % 2 == 0 else FILL_EVEN
    align = GLOBAL_ALIGNMENT_CENTER if horizontal == 'center' else GLOBAL_ALIGNMENT_RIGHT
    return _wo_cell(
        ws, value,
        font=FONT_DATA,
        fill=row_type_fill,
        alignment=align,
        border=GLOBAL_BORDER,
        number_format=number_format,
    )


def _wo_balance_cell(ws, value):
    """ساخت یک WriteOnlyCell با استایل شرطی موازنه (جایگزین _apply_balance_style در write-only mode)."""
    if isinstance(value, str) and "در دست بررسی" in value:
        return _wo_cell(
            ws, value,
            font=Font(name='Calibri', size=10, bold=True, color="595959"),
            fill=PatternFill(fill_type='solid', fgColor="F2F2F2"),
            alignment=GLOBAL_ALIGNMENT_CENTER,
            border=GLOBAL_BORDER,
        )
    fval = float(value)
    if fval > 0:
        return _wo_cell(ws, fval, font=FONT_POS, fill=FILL_POS, alignment=GLOBAL_ALIGNMENT_CENTER, border=GLOBAL_BORDER, number_format='#,##0.00')
    elif fval < 0:
        return _wo_cell(ws, fval, font=FONT_NEG, fill=FILL_NEG, alignment=GLOBAL_ALIGNMENT_CENTER, border=GLOBAL_BORDER, number_format='#,##0.00')
    else:
        return _wo_cell(ws, fval, font=FONT_ZERO, fill=FILL_ZERO, alignment=GLOBAL_ALIGNMENT_CENTER, border=GLOBAL_BORDER, number_format='#,##0.00')


def _wo_status_cell(ws, balance_val, balance_label):
    """ساخت یک WriteOnlyCell برای ستون وضعیت نهایی."""
    if isinstance(balance_val, str):
        return _wo_cell(ws, balance_label, font=Font(name='Calibri', size=10, bold=True, color="595959"),
                        fill=PatternFill(fill_type='solid', fgColor="F2F2F2"), alignment=GLOBAL_ALIGNMENT_CENTER, border=GLOBAL_BORDER)
    elif balance_val > 0:
        return _wo_cell(ws, balance_label, font=FONT_POS, fill=FILL_POS, alignment=GLOBAL_ALIGNMENT_CENTER, border=GLOBAL_BORDER)
    elif balance_val < 0:
        return _wo_cell(ws, balance_label, font=FONT_NEG, fill=FILL_NEG, alignment=GLOBAL_ALIGNMENT_CENTER, border=GLOBAL_BORDER)
    else:
        return _wo_cell(ws, balance_label, font=FONT_ZERO, fill=FILL_ZERO, alignment=GLOBAL_ALIGNMENT_CENTER, border=GLOBAL_BORDER)


def generate_global_material_balance_excel(is_superuser: bool = False, task_id=None, resume_from=None) -> bytes:
    """
    تولید گزارش موازنه متریال کل کارگاه به صورت تجمیعی با xlsxwriter
    فاز‌بندی پیشرفت: DB(0-20%) → Excel(20-95%) → Upload(95-100%)
    """
    import xlsxwriter
    import time
    from .models import ExportTask
    from .tasks import send_task_progress

    start_time = time.time()

    # ── فاز ۱: خواندن دیتابیس (0-20%) ──
    if task_id:
        ExportTask.objects.filter(pk=task_id).update(progress=1, eta=0)
        send_task_progress(task_id, 1, phase='در حال خواندن دیتابیس...', eta=None)

    rows = get_global_material_balance_rows_data()

    db_elapsed = time.time() - start_time
    if task_id:
        ExportTask.objects.filter(pk=task_id).update(progress=20)
        total_eta = int(db_elapsed * 4)  # rough estimate: DB ~ 25% of total time
        send_task_progress(task_id, 20, phase='در حال ساخت فایل اکسل...', eta=total_eta)

    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output, {'constant_memory': True})
    ws = wb.add_worksheet("موازنه کل")
    
    if not rows:
        error_format = wb.add_format({'font_name': 'Calibri', 'font_size': 12, 'bold': True, 'font_color': '#9C0006'})
        ws.write(0, 0, "هیچ داده‌ای برای گزارش‌گیری وجود ندارد.", error_format)
        wb.close()
        return output.getvalue()

    # ── فاز ۲: نوشتن ردیف‌های اکسل (20-95%) ──
    _build_global_sheet(ws, rows, wb, task_id=task_id, start_row_idx=0, phase_start_time=time.time())

    # ── فاز ۳: بسته‌بندی فایل (95-100%) ──
    if task_id:
        send_task_progress(task_id, 95, phase='در حال نهایی‌سازی فایل...', eta=2)

    wb.close()
    output.seek(0)
    return output.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# تابع کمکی: دانلود مستقیم از Django View
# ─────────────────────────────────────────────────────────────────────────────
def get_balance_excel_response(contractor_id=None, material_id=None, is_superuser=False):
    """
    یک HttpResponse آماده برای دانلود فایل اکسل از Django View برمی‌گرداند.

    مثال استفاده در views.py:
        from .services import get_balance_excel_response

        def download_balance(request):
            contractor_id = request.GET.get('contractor')
            return get_balance_excel_response(contractor_id=contractor_id, is_superuser=request.user.is_superuser)
    """
    from django.http import HttpResponse

    file_bytes = generate_material_balance_excel(
        contractor_id=contractor_id,
        material_id=material_id,
        is_superuser=is_superuser,
    )
    filename = f"material_balance_{jdatetime.date.today().isoformat()}.xlsx"
    response = HttpResponse(
        content=file_bytes,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def get_global_balance_excel_response(is_superuser=False):
    """
    دانلود گزارش کلی موازنه.
    """
    from django.http import HttpResponse

    file_bytes = generate_global_material_balance_excel(is_superuser=is_superuser)
    filename = f"global_material_balance_{jdatetime.date.today().isoformat()}.xlsx"
    response = HttpResponse(
        content=file_bytes,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────────────────────────────────────────────────────────
# تابع خروجی JSON برای API داشبورد
# ─────────────────────────────────────────────────────────────────────────────
def get_contractors_balance_summary() -> list[dict]:
    """
    محاسبه موازنه کلی تمامی پیمانکاران برای نمایش در داشبورد یا API.
    برگرداندن لیستی از دیکشنری‌ها.
    """
    from .models import WarehouseTransaction, TechnicalOfficeApproval, MaterialItem, Contractor
    from django.db.models import Sum

    # تراکنش‌های خروجی
    issue_qs = WarehouseTransaction.objects.filter(transaction_type='OUT')
    issued_aggs = issue_qs.values('contractor_id', 'material_id', 'contract_number', 'contract_subject').annotate(total_qty=Sum('quantity'))
    issued_map: dict[tuple, Decimal] = {}
    for agg in issued_aggs:
        k = (agg['contractor_id'], agg['material_id'], agg['contract_number'] or '', agg['contract_subject'] or '')
        issued_map[k] = agg['total_qty'] or Decimal('0')

    # تاییدیه‌ها
    approval_qs = TechnicalOfficeApproval.objects.all()
    appr_aggs = approval_qs.values('contractor_id', 'material_id', 'contract_number', 'contract_subject').annotate(total_appr=Sum('approved_quantity'))
    approved_map: dict[tuple, Decimal] = {}
    for agg in appr_aggs:
        k = (agg['contractor_id'], agg['material_id'], agg['contract_number'] or '', agg['contract_subject'] or '')
        approved_map[k] = agg['total_appr'] or Decimal('0')

    # کلیدها و مدل‌ها
    all_keys = set(issued_map.keys()) | set(approved_map.keys())
    contractor_ids = {k[0] for k in all_keys if k[0] is not None}
    material_ids   = {k[1] for k in all_keys}
    
    contractors = {c.id: c for c in Contractor.objects.filter(id__in=contractor_ids)}
    materials = {m.id: m for m in MaterialItem.objects.filter(id__in=material_ids)}

    # محاسبه موازنه برای هر ترکیب (پیمانکار، متریال) و جمع زدن بر اساس پیمانکار به تفکیک واحد
    contractor_balances = {c_id: {} for c_id in contractor_ids}
    contractor_under_review = {c_id: 0 for c_id in contractor_ids}

    for c_id, m_id, c_num, c_subj in all_keys:
        material = materials.get(m_id)
        contractor = contractors.get(c_id)
        if not material or not contractor:
            continue

        total_issued  = issued_map.get((c_id, m_id, c_num, c_subj), Decimal('0'))
        
        if (c_id, m_id, c_num, c_subj) not in approved_map:
            # اگر تاییدیه‌ای وجود ندارد، در دست بررسی است و در مجموع بالانس تاثیر نمی‌دهد
            contractor_under_review[c_id] += 1
        else:
            approved_work = approved_map[(c_id, m_id, c_num, c_subj)]
            waste_pct     = material.waste_percentage / Decimal('100')
            allowed_waste = (approved_work * waste_pct).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            balance       = (total_issued - (approved_work + allowed_waste)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            
            unit = material.unit
            if unit not in contractor_balances[c_id]:
                contractor_balances[c_id][unit] = Decimal('0')
            contractor_balances[c_id][unit] += balance

    # ساخت خروجی نهایی به شکل لیست
    summary = []
    for c_id, unit_bals in contractor_balances.items():
        contractor = contractors.get(c_id)
        balances_json = {unit: float(val) for unit, val in unit_bals.items()}
        summary.append({
            'contractor_id': c_id,
            'contractor_name': contractor.get_full_name() if contractor else "ناشناس",
            'balances': balances_json,
            'under_review_count': contractor_under_review[c_id],
        })

    # مرتب‌سازی بر اساس نام پیمانکار
    summary.sort(key=lambda x: x['contractor_name'])
    return summary


# ─────────────────────────────────────────────────────────────────────────────
# خروجی اکسل انباردار
# ─────────────────────────────────────────────────────────────────────────────
def generate_warehouse_inventory_excel(is_superuser: bool = False) -> bytes:
    from .models import WarehouseTransaction
    qs = WarehouseTransaction.objects.select_related('contractor', 'material', 'material__work_category').defer('bill_of_lading_image', 'exit_document_image').order_by('-created_at')

    wb = Workbook()
    _register_named_styles(wb)
    ws = wb.active
    ws.title = "تراکنش‌های انبار"
    ws.sheet_view.rightToLeft = True

    TOTAL_COLS = 14
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
    _apply_header_style(ws.cell(1, 1), "لیست تراکنش‌های انبار - شرکت جهانپارس", font_size=16)
    ws.row_dimensions[1].height = 38

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=TOTAL_COLS)
    sub = ws.cell(2, 1)
    sub.value = f"تاریخ گزارش: {jdatetime.date.today().isoformat()}"
    sub.font = Font(name='Calibri', size=10)
    sub.fill = PatternFill(fill_type='solid', fgColor=COLOR_INFO_BG)
    sub.alignment = _rtl_alignment(horizontal='center')
    ws.row_dimensions[2].height = 20

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=TOTAL_COLS)
    ws.cell(3, 1).fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
    ws.row_dimensions[3].height = 6

    HEADERS = ["ردیف", "نوع تراکنش", "تاریخ", "پیمانکار", "شماره قرارداد", "موضوع قرارداد", "رسته کاری", "کالا", "سایز", "جنس", "ضخامت", "تعداد/مقدار", "واحد", "توضیحات"]
    COL_WIDTHS = [6, 12, 14, 20, 16, 25, 16, 24, 12, 14, 12, 14, 10, 30]

    for i, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        c = ws.cell(4, i)
        _apply_header_style(c, h, font_size=10, bg_color=COLOR_SUB_HEADER_BG)
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[4].height = 28

    last_row = 4
    for row_idx, txn in enumerate(qs.iterator(chunk_size=2000)):
        r = row_idx + 5
        last_row = r
        _apply_data_style(ws.cell(r, 1), row_idx + 1, row_idx, horizontal='center')
        
        type_cell = ws.cell(r, 2)
        type_cell.value = "ورود" if txn.transaction_type == 'IN' else "خروج"
        type_cell.style = "data_in" if txn.transaction_type == 'IN' else "data_out"

        _apply_data_style(ws.cell(r, 3), str(txn.date), row_idx, horizontal='center')
        _apply_data_style(ws.cell(r, 4), txn.contractor.get_full_name() if txn.contractor else "—", row_idx)
        _apply_data_style(ws.cell(r, 5), txn.contract_number or "—", row_idx, horizontal='center')
        _apply_data_style(ws.cell(r, 6), txn.contract_subject or "—", row_idx)
        _apply_data_style(ws.cell(r, 7), txn.material.work_category.name if txn.material and txn.material.work_category else "—", row_idx)
        _apply_data_style(ws.cell(r, 8), txn.material.name if txn.material else "—", row_idx)
        _apply_data_style(ws.cell(r, 9), txn.material.size if txn.material and txn.material.size else "—", row_idx, horizontal='center')
        _apply_data_style(ws.cell(r, 10), txn.material.material_type if txn.material and txn.material.material_type else "—", row_idx, horizontal='center')
        _apply_data_style(ws.cell(r, 11), txn.material.thickness if txn.material and txn.material.thickness else "—", row_idx, horizontal='center')
        _apply_data_style(ws.cell(r, 12), float(txn.quantity), row_idx, number_format='#,##0.00')
        _apply_data_style(ws.cell(r, 13), txn.material.get_unit_display() if txn.material else "—", row_idx, horizontal='center')
        _apply_data_style(ws.cell(r, 14), txn.bill_of_lading or "", row_idx)

    ws.auto_filter.ref = f"A4:N{last_row}"
    ws.freeze_panes = "A5"

    if not is_superuser:
        ws.protection.sheet = True
        ws.protection.password = "jahanpars2026"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def get_warehouse_inventory_excel_response(is_superuser: bool = False):
    from django.http import HttpResponse
    file_bytes = generate_warehouse_inventory_excel(is_superuser=is_superuser)
    filename = f"warehouse_inventory_{jdatetime.date.today().isoformat()}.xlsx"
    response = HttpResponse(
        content=file_bytes,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────────────────────────────────────────────────────────
# خروجی اکسل لیست پیمانکاران
# ─────────────────────────────────────────────────────────────────────────────
def generate_contractors_list_excel(is_superuser: bool = False) -> bytes:
    from .models import Contractor
    qs = Contractor.objects.all().order_by('first_name', 'last_name')

    wb = Workbook()
    ws = wb.active
    ws.title = "لیست پیمانکاران"
    ws.sheet_view.rightToLeft = True

    TOTAL_COLS = 3
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
    _apply_header_style(ws.cell(1, 1), "لیست پیمانکاران - شرکت جهانپارس", font_size=16)
    ws.row_dimensions[1].height = 38

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=TOTAL_COLS)
    sub = ws.cell(2, 1)
    sub.value = f"تاریخ گزارش: {jdatetime.date.today().isoformat()}"
    sub.font = Font(name='Calibri', size=10)
    sub.fill = PatternFill(fill_type='solid', fgColor=COLOR_INFO_BG)
    sub.alignment = _rtl_alignment(horizontal='center')
    ws.row_dimensions[2].height = 20

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=TOTAL_COLS)
    ws.cell(3, 1).fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
    ws.row_dimensions[3].height = 6

    HEADERS = ["ردیف", "نام", "نام خانوادگی"]
    COL_WIDTHS = [10, 30, 30]

    for i, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        c = ws.cell(4, i)
        _apply_header_style(c, h, font_size=10, bg_color=COLOR_SUB_HEADER_BG)
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[4].height = 28

    last_row = 4
    for row_idx, contractor in enumerate(qs.iterator(chunk_size=2000)):
        r = row_idx + 5
        last_row = r
        _apply_data_style(ws.cell(r, 1), row_idx + 1, row_idx, horizontal='center')
        _apply_data_style(ws.cell(r, 2), contractor.first_name, row_idx)
        _apply_data_style(ws.cell(r, 3), contractor.last_name, row_idx)

    ws.auto_filter.ref = f"A4:C{last_row}"
    ws.freeze_panes = "A5"

    if not is_superuser:
        ws.protection.sheet = True
        ws.protection.password = "jahanpars2026"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def get_contractors_excel_response(is_superuser: bool = False):
    from django.http import HttpResponse
    file_bytes = generate_contractors_list_excel(is_superuser=is_superuser)
    filename = f"contractors_list_{jdatetime.date.today().isoformat()}.xlsx"
    response = HttpResponse(
        content=file_bytes,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────────────────────────────────────────────────────────
# خروجی اکسل لیست تاییدیه‌ها
# ─────────────────────────────────────────────────────────────────────────────
def generate_approvals_list_excel(is_superuser: bool = False) -> bytes:
    from .models import TechnicalOfficeApproval
    qs = TechnicalOfficeApproval.objects.select_related('contractor', 'material').order_by('-created_at')

    wb = Workbook()
    _register_named_styles(wb)
    ws = wb.active
    ws.title = "لیست تاییدیه‌های دفتر فنی"
    ws.sheet_view.rightToLeft = True

    TOTAL_COLS = 12
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
    _apply_header_style(ws.cell(1, 1), "لیست تاییدیه‌های دفتر فنی - شرکت جهانپارس", font_size=16)
    ws.row_dimensions[1].height = 38

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=TOTAL_COLS)
    sub = ws.cell(2, 1)
    sub.value = f"تاریخ گزارش: {jdatetime.date.today().isoformat()}"
    sub.font = Font(name='Calibri', size=10)
    sub.fill = PatternFill(fill_type='solid', fgColor=COLOR_INFO_BG)
    sub.alignment = _rtl_alignment(horizontal='center')
    ws.row_dimensions[2].height = 20

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=TOTAL_COLS)
    ws.cell(3, 1).fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
    ws.row_dimensions[3].height = 6

    HEADERS = ["ردیف", "تاریخ تایید", "پیمانکار", "نام کالا", "سایز", "جنس", "ضخامت", "واحد", "موضوع / توضیحات", "مقدار تایید شده", "شماره قرارداد", "توضیحات تکمیلی"]
    COL_WIDTHS = [6, 14, 25, 24, 12, 14, 12, 10, 30, 15, 15, 30]

    for i, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        c = ws.cell(4, i)
        _apply_header_style(c, h, font_size=10, bg_color=COLOR_SUB_HEADER_BG)
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[4].height = 28

    last_row = 4
    for row_idx, approval in enumerate(qs.iterator(chunk_size=2000)):
        r = row_idx + 5
        last_row = r
        _apply_data_style(ws.cell(r, 1), row_idx + 1, row_idx, horizontal='center')
        _apply_data_style(ws.cell(r, 2), str(approval.approval_date), row_idx, horizontal='center')
        _apply_data_style(ws.cell(r, 3), approval.contractor.get_full_name() if approval.contractor else "—", row_idx)
        _apply_data_style(ws.cell(r, 4), approval.material.name if approval.material else "—", row_idx)
        _apply_data_style(ws.cell(r, 5), approval.material.size if approval.material and approval.material.size else "—", row_idx, horizontal='center')
        _apply_data_style(ws.cell(r, 6), approval.material.material_type if approval.material and approval.material.material_type else "—", row_idx, horizontal='center')
        _apply_data_style(ws.cell(r, 7), approval.material.thickness if approval.material and approval.material.thickness else "—", row_idx, horizontal='center')
        _apply_data_style(ws.cell(r, 8), approval.material.get_unit_display() if approval.material else "—", row_idx, horizontal='center')
        _apply_data_style(ws.cell(r, 9), approval.contract_subject or "—", row_idx)
        _apply_data_style(ws.cell(r, 10), float(approval.approved_quantity), row_idx, number_format='#,##0.00')
        _apply_data_style(ws.cell(r, 11), approval.contract_number or "—", row_idx, horizontal='center')
        _apply_data_style(ws.cell(r, 12), "—", row_idx)

    ws.auto_filter.ref = f"A4:L{last_row}"
    ws.freeze_panes = "A5"

    if not is_superuser:
        ws.protection.sheet = True
        ws.protection.password = "jahanpars2026"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def get_approvals_excel_response(is_superuser: bool = False):
    from django.http import HttpResponse
    file_bytes = generate_approvals_list_excel(is_superuser=is_superuser)
    filename = f"approvals_list_{jdatetime.date.today().isoformat()}.xlsx"
    response = HttpResponse(
        content=file_bytes,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
